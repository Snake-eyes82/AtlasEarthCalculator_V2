# main.py

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys

# Import our custom modules
import constants
import utils
from widgets import IntegerEntry
from current_earnings_calculator import CurrentEarningsCalculator
from goal_calculator import GoalCalculator
from next_tier_calculator import NextTierCalculator
from custom_tier_calculator import CustomTierCalculator

# External libraries for export (will need to be installed)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    openpyxl = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None


# Custom Entry widget for integer input with robust validation and cleaning
class IntegerEntry(ttk.Entry):
    def __init__(self, master=None, **kwargs):
        # Use a StringVar that we control
        self.var = kwargs.pop('textvariable', tk.StringVar(value="0"))
        super().__init__(master, textvariable=self.var, **kwargs)

        # Register validation command to allow only digits
        vcmd = self.register(self._validate_input)
        self.config(validate="key", validatecommand=(vcmd, '%P')) # %P is the new value of the entry

        # Bind to clean and update on focus out or Return key
        self.bind("<FocusOut>", self._clean_and_update)
        self.bind("<Return>", self._clean_and_update)

        # Initial cleaning in case default value isn't clean (e.g., if it was "00")
        self._clean_and_update()

    def _validate_input(self, new_value):
        """Allows only digits and empty string."""
        if new_value == "":
            return True
        return new_value.isdigit()

    def _clean_and_update(self, event=None):
        """
        Cleans the input (removes leading zeros, ensures non-negative integer),
        and then triggers the main application's update_all_calculations.
        """
        current_value = self.var.get()
        
        if current_value.strip() == "":
            cleaned_value = "0"
        else:
            try:
                # Convert to integer to handle leading zeros automatically (e.g., "045" -> 45)
                int_val = int(current_value)
                if int_val < 0: # Ensure non-negative
                    int_val = 0
                cleaned_value = str(int_val) # Convert back to string (e.g., 45 -> "45")
            except ValueError:
                # This should ideally not happen if _validate_input is working,
                # but as a fallback for non-digit input (e.g., paste)
                cleaned_value = "0"
        
        # Only update the StringVar if the value actually changed.
        # This is crucial to prevent infinite loops with trace_add if it were used,
        # and generally good practice to avoid unnecessary widget updates.
        if self.var.get() != cleaned_value:
            self.var.set(cleaned_value)
        
        # Trigger the main application's update.
        # We access the root window's 'app_instance' attribute.
        app_instance = self.master.winfo_toplevel().app_instance

        # IMPORTANT FIX: Only call update_all_calculations if the main app and its calculators are fully initialized.
        # This prevents the AttributeError during startup.
        if hasattr(app_instance, 'current_earnings_calculator') and \
           hasattr(app_instance, 'goal_calculator') and \
           hasattr(app_instance, 'next_tier_calculator'):
            app_instance.update_all_calculations()


class AtlasEarthApp:
    def __init__(self, master):
        self.master = master
        master.title("Atlas Earth Calculator")
        master.geometry("800x700")
        master.resizable(True, True)

        # Store a reference to this app instance in the root window.
        # This allows our custom IntegerEntry widgets to call update_all_calculations.
        master.app_instance = self 

        self.style = ttk.Style()
        self.style.configure("TLabel", font=("Helvetica", 10))
        self.style.configure("TButton", font=("Helvetica", 10))
        self.style.configure("TEntry", font=("Helvetica", 10))
        self.style.configure("TCheckbutton", font=("Helvetica", 10))
        self.style.configure("TRadiobutton", font=("Helvetica", 10))
        self.style.configure("TCombobox", font=("Helvetica", 10))
        self.style.configure("TLabelframe.Label", font=("Helvetica", 12, "bold"))

        self.output_font_fixed = ("Courier", 10, "bold")

        # --- Tkinter Variables for Main Inputs ---
        self.parcel_vars = {p_type: tk.StringVar(value="0") for p_type in constants.PARCEL_RATES_PER_SECOND.keys()}
        self.total_parcels_var = tk.StringVar(value="0") # To display total parcels
        self.badge_count_var = tk.StringVar(value="0")
        self.boost_hours_var = tk.StringVar(value="0") # CORRECTED: Default to 0 as requested
        self.srb_boost_enabled = tk.BooleanVar(value=False)
        self.fictive_badge_boost_enabled = tk.BooleanVar(value=False)
        self.fictive_badge_boost_percent_var = tk.StringVar(value="0.0")
        self.selected_region_var = tk.StringVar(value="United States")

        # --- Setup Menu Bar ---
        self._create_menu_bar()

        # --- Setup Scrollable Frame ---
        self.main_canvas = tk.Canvas(master)
        self.main_canvas.pack(side="left", fill="both", expand=True)

        self.scrollbar = ttk.Scrollbar(master, orient="vertical", command=self.main_canvas.yview)
        self.scrollbar.pack(side="right", fill="y")

        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        # Bind to configure event for the canvas to update scrollregion when canvas size changes
        self.main_canvas.bind('<Configure>', lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        
        self.content_frame = ttk.Frame(self.main_canvas)
        self.main_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")
        # Bind to configure event for the content_frame to update scrollregion when content size changes
        self.content_frame.bind('<Configure>', lambda e: self.main_canvas.configure(scrollregion=self.content_frame.bbox("all")))
        self.main_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # --- Input Section ---
        input_frame = ttk.LabelFrame(self.content_frame, text="Your Parcel and Boost Info")
        input_frame.pack(padx=10, pady=5, fill="x")

        # Regional Selection Dropdown
        regional_frame = ttk.Frame(input_frame)
        regional_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(regional_frame, text="Select Region:").pack(side="left", padx=5)
        
        regions = list(constants.REGIONAL_AD_BOOST_DATA.keys())
        self.region_combobox = ttk.Combobox(regional_frame, textvariable=self.selected_region_var, values=regions, state="readonly")
        self.region_combobox.pack(side="left", padx=5)
        self.region_combobox.bind("<<ComboboxSelected>>", lambda event: self.update_all_calculations())

        # Parcel Inputs - NOW USING IntegerEntry for robust input handling
        for p_type in constants.PARCEL_RATES_PER_SECOND.keys():
            row_frame = ttk.Frame(input_frame)
            row_frame.pack(fill="x", padx=5, pady=5)
            ttk.Label(row_frame, text=f"{p_type.capitalize()} Parcels:").pack(side="left", padx=5)
            # Use our custom IntegerEntry here, passing its textvariable
            IntegerEntry(row_frame, width=10, textvariable=self.parcel_vars[p_type]).pack(side="left", padx=5)

        # Display Total Parcels
        total_parcels_frame = ttk.Frame(input_frame)
        total_parcels_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(total_parcels_frame, text="Total Parcels:").pack(side="left", padx=5)
        ttk.Label(total_parcels_frame, textvariable=self.total_parcels_var, font=("Helvetica", 10, "bold")).pack(side="left", padx=5)

        # Ad Boost Hours Input - NOW USING IntegerEntry for consistency and robust input
        boost_frame = ttk.Frame(input_frame)
        boost_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(boost_frame, text="Ad Boost Hours/Day:").pack(side="left", padx=5)
        # Use our custom IntegerEntry here
        IntegerEntry(boost_frame, width=5, textvariable=self.boost_hours_var).pack(side="left", padx=5)

        # Super Rent Boost Checkbox
        srb_check = ttk.Checkbutton(boost_frame, text=f"Force Super Rent Boost ({constants.SUPER_RENT_BOOST_MULTIPLIER}x)",
                                    variable=self.srb_boost_enabled,
                                    command=self.update_all_calculations)
        srb_check.pack(side="left", padx=10)

        # Badges Owned Input and Fictive Badge Boost - NOW USING IntegerEntry for Badges Owned
        badge_frame = ttk.LabelFrame(input_frame, text="Badge Info")
        badge_frame.pack(padx=5, pady=5, fill="x")
        ttk.Label(badge_frame, text="Badges Owned:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        # Use our custom IntegerEntry here
        IntegerEntry(badge_frame, width=5, textvariable=self.badge_count_var).grid(row=0, column=1, sticky="w", padx=5, pady=2)

        fictive_boost_check = ttk.Checkbutton(badge_frame, text="Use Fictive Badge Boost (%)",
                                            variable=self.fictive_badge_boost_enabled,
                                            command=self._toggle_fictive_badge_boost)
        fictive_boost_check.grid(row=1, column=0, sticky="w", padx=5, pady=2)
        # Fictive Badge Boost Entry - this is a float, so keep as ttk.Entry for now
        self.fictive_badge_boost_entry = ttk.Entry(badge_frame, width=8, textvariable=self.fictive_badge_boost_percent_var, state="disabled")
        self.fictive_badge_boost_entry.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        # This one still needs KeyRelease bind for calculation update as it's not an IntegerEntry
        self.fictive_badge_boost_entry.bind("<KeyRelease>", lambda event: self.update_all_calculations())


        # --- Tabbed Interface ---
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(pady=10, expand=True, fill="both")

        # Current Earnings Tab
        self.current_earnings_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.current_earnings_tab, text="Current Earnings")
        self.current_earnings_calculator = CurrentEarningsCalculator(self.current_earnings_tab, self.get_user_inputs)

        # Goal Calculator Tab
        self.goal_calculator_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.goal_calculator_tab, text="Parcels for Goal")
        self.goal_calculator = GoalCalculator(self.goal_calculator_tab, self.get_user_inputs)

        # Next Tier Tab
        self.next_tier_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.next_tier_tab, text="Next Tier Info")
        self.next_tier_calculator = NextTierCalculator(self.next_tier_tab, self.get_user_inputs)

        # Placeholder Tabs
        self.custom_tier_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.custom_tier_tab, text="Custom Tier")
        self.custom_tier_calculator = CustomTierCalculator(self.custom_tier_tab, self.get_user_inputs) # Instantiate it

        self.additional_info_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.additional_info_tab, text="Additional Info")

        self.srb_event_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.srb_event_tab, text="SRB Event")
        
        # --- Buttons ---
        button_frame = ttk.Frame(self.content_frame)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Clear All", command=self.clear_all).pack(side="left", padx=5)

        # Initial calculation update
        self.update_all_calculations()

    def _on_mousewheel(self, event):
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def _create_menu_bar(self):
        menubar = tk.Menu(self.master)
        self.master.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Export to XLSX", command=self._export_to_xlsx, state="normal" if openpyxl else "disabled")
        file_menu.add_command(label="Export to CSV", command=self._export_to_csv)
        file_menu.add_command(label="Export to PDF", command=self._export_to_pdf, state="normal" if FPDF else "disabled")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.master.quit)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About Calculator", command=self._show_about_dialog)

        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Settings", menu=settings_menu)
        settings_menu.add_command(label="API/AI Settings (Placeholder)", command=self._open_settings_dialog)

    def _show_about_dialog(self):
        about_text = (
            "Atlas Earth Calculator\n"
            "Version: 1.0 (Modular)\n"
            "Developed by: Your Name / Gemini AI\n"
            "Special Thanks: Jason (for detailed requirements and testing)\n\n"
            "This calculator helps Atlas Earth players estimate earnings and plan parcel acquisitions.\n"
            "All data is based on publicly available information and user-provided charts."
        )
        messagebox.showinfo("About Atlas Earth Calculator", about_text)

    def _open_settings_dialog(self):
        messagebox.showinfo("Settings", "API/AI Settings will be implemented here later.")

    def _toggle_fictive_badge_boost(self):
        if self.fictive_badge_boost_enabled.get():
            self.fictive_badge_boost_entry.config(state="normal")
        else:
            self.fictive_badge_boost_entry.config(state="disabled")
            self.fictive_badge_boost_percent_var.set("0.0")
        self.update_all_calculations()

    def get_user_inputs(self):
        """Retrieves and validates all user inputs from the GUI."""
        parcels = {p_type: 0 for p_type in constants.PARCEL_RATES_PER_SECOND.keys()}
        total_parcels = 0
        # IntegerEntry now handles its own validation and cleaning,
        # so we can simply get the value and convert to int.
        for p_type, parcel_var in self.parcel_vars.items():
            try:
                count = int(parcel_var.get())
                if count < 0: # Should be prevented by IntegerEntry, but as fallback
                    count = 0
                parcels[p_type] = count
                total_parcels += count
            except ValueError:
                # This error should ideally not be hit if IntegerEntry is working correctly
                messagebox.showerror("Input Error", f"Invalid input for {p_type.capitalize()} parcels. Please enter a whole number.")
                return None

        # Update the total_parcels_var for display
        self.total_parcels_var.set(str(total_parcels)) # NEW LINE HERE

        # boost_hours_var is now handled by IntegerEntry as well
        try:
            boost_hours_str = self.boost_hours_var.get()
            boost_hours = float(boost_hours_str) # Use float for boost hours
            if not (0 <= boost_hours <= 24): # Range validation still needed here
                messagebox.showerror("Input Error", "Ad Boost Hours must be between 0 and 24.")
                return None
        except ValueError:
            messagebox.showerror("Input Error", "Ad Boost Hours must be a number.")
            return None

        # badge_count_var is now handled by IntegerEntry as well
        try:
            badge_count = int(self.badge_count_var.get())
            if badge_count < 0: # Should be prevented by IntegerEntry, but as fallback
                badge_count = 0
        except ValueError:
            messagebox.showerror("Input Error", "Badges Owned must be a non-negative whole number.")
            return None

        # Fictive Badge Boost is still a regular Entry as it can be a float
        fictive_badge_boost_percent = 0.0
        if self.fictive_badge_boost_enabled.get():
            try:
                fictive_badge_boost_percent_str = self.fictive_badge_boost_percent_var.get()
                fictive_badge_boost_percent = float(fictive_badge_boost_percent_str) if fictive_badge_boost_percent_str.strip() != "" else 0.0
                if fictive_badge_boost_percent > 1.0: # Assume user might enter 5 for 5%
                    fictive_badge_boost_percent /= 100.0
                if not (0.0 <= fictive_badge_boost_percent <= 1.0):
                    messagebox.showerror("Input Error", "Fictive Badge Boost must be between 0.0 and 100.0 (e.g., 5 for 5% or 0.05 for 5%).")
                    return None
            except ValueError:
                messagebox.showerror("Input Error", "Fictive Badge Boost must be a number.")
                return None

        return {
            "parcels": parcels,
            "total_parcels": total_parcels,
            "boost_hours": boost_hours,
            "badge_count": badge_count,
            "srb_boost_enabled": self.srb_boost_enabled.get(),
            "fictive_badge_boost_enabled": self.fictive_badge_boost_enabled.get(),
            "fictive_badge_boost_percent": fictive_badge_boost_percent,
            "selected_region": self.selected_region_var.get(),
        }

    def update_all_calculations(self):
        """Triggers updates in all calculator modules."""
        # Only update if the calculator objects have been initialized
        if hasattr(self, 'current_earnings_calculator'):
            self.current_earnings_calculator.update_display()
        # Goal calculator does not auto-update, it's triggered by its own button.
        if hasattr(self, 'next_tier_calculator'):
            self.next_tier_calculator.update_display()
        if hasattr(self, 'custom_tier_calculator'): # Update the new custom tier tab
            self.custom_tier_calculator.update_display()

    def _get_all_calculated_data(self):
        """Collects all relevant calculated data from the calculator modules."""
        data = {}
        
        # Get data from CurrentEarningsCalculator
        data["current_earnings"] = self.current_earnings_calculator.get_export_data()
        
        # Get data from GoalCalculator (last calculated goal)
        data["goal_data"] = self.goal_calculator.get_export_data()

        # Get data from NextTierCalculator
        data["next_tier_info"] = self.next_tier_calculator.get_export_data()

        # Add main input data for context
        data["user_inputs"] = {
            "common_parcels": self.parcel_vars["common"].get(),
            "rare_parcels": self.parcel_vars["rare"].get(),
            "epic_parcels": self.parcel_vars["epic"].get(),
            "legendary_parcels": self.parcel_vars["legendary"].get(),
            # Summing values from StringVars which are now guaranteed to be clean digits by IntegerEntry
            "total_parcels_input": sum(int(self.parcel_vars[p].get()) for p in self.parcel_vars),
            "badge_count": self.badge_count_var.get(),
            "ad_boost_hours_day": self.boost_hours_var.get(),
            "force_srb": self.srb_boost_enabled.get(),
            "fictive_badge_boost_enabled": self.fictive_badge_boost_enabled.get(),
            "fictive_badge_boost_percent": self.fictive_badge_boost_percent_var.get(),
            "selected_region": self.selected_region_var.get(),
        }
        return data

    def _export_to_xlsx(self):
        if not openpyxl:
            messagebox.showerror("Error", "openpyxl library not found. Please install it using 'pip install openpyxl'")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                title="Save Calculations to Excel")
        if not file_path:
            return

        try:
            data = self._get_all_calculated_data()
            workbook = openpyxl.Workbook()
            
            # --- Input Data Sheet ---
            input_sheet = workbook.active
            input_sheet.title = "User Inputs"
            input_sheet.append(["Input Type", "Value"])
            for key, value in data["user_inputs"].items():
                input_sheet.append([key.replace('_', ' ').title(), value])
            
            # --- Current Earnings Sheet ---
            current_earnings_sheet = workbook.create_sheet("Current Earnings")
            current_earnings_sheet.append(["Timeframe", "Base Earnings", "With Ad Boost"])
            for tf, values in data["current_earnings"].items():
                current_earnings_sheet.append([f"Per {tf.capitalize()}:", values['base'], values['boosted']])

            # --- Goal Data Sheet ---
            goal_sheet = workbook.create_sheet("Parcels for Goal")
            goal_sheet.append(["Goal Type", "Value"])
            for key, value in data["goal_data"].items():
                if isinstance(value, dict): # For breakdown
                    goal_sheet.append([key.replace('_', ' ').title(), ""])
                    for sub_key, sub_value in value.items():
                        goal_sheet.append([f"  {sub_key.replace('_', ' ').title()}", sub_value])
                else:
                    goal_sheet.append([key.replace('_', ' ').title(), value])

            # --- Next Tier Info Sheet ---
            next_tier_sheet = workbook.create_sheet("Next Tier Info")
            next_tier_sheet.append(["Info Type", "Value"])
            for key, value in data["next_tier_info"].items():
                next_tier_sheet.append([key.replace('_', ' ').title(), value])

            workbook.save(file_path)
            messagebox.showinfo("Export Success", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to XLSX: {e}")

    def _export_to_csv(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".csv",
                                                filetypes=[("CSV files", "*.csv")],
                                                title="Save Calculations to CSV")
        if not file_path:
            return

        try:
            data = self._get_all_calculated_data()
            with open(file_path, 'w', newline='') as csvfile:
                import csv
                writer = csv.writer(csvfile)

                # User Inputs
                writer.writerow(["--- User Inputs ---"])
                writer.writerow(["Input Type", "Value"])
                for key, value in data["user_inputs"].items():
                    writer.writerow([key.replace('_', ' ').title(), value])
                writer.writerow([])

                # Current Earnings
                writer.writerow(["--- Current Earnings ---"])
                writer.writerow(["Timeframe", "Base Earnings", "With Ad Boost"])
                for tf, values in data["current_earnings"].items():
                    writer.writerow([f"Per {tf.capitalize()}:", values['base'], values['boosted']])
                writer.writerow([])

                # Goal Data
                writer.writerow(["--- Parcels for Goal ---"])
                writer.writerow(["Goal Type", "Value"])
                for key, value in data["goal_data"].items():
                    if isinstance(value, dict):
                        writer.writerow([key.replace('_', ' ').title(), ""])
                        for sub_key, sub_value in value.items():
                            writer.writerow([f"  {sub_key.replace('_', ' ').title()}", sub_value])
                    else:
                        writer.writerow([key.replace('_', ' ').title(), value])
                writer.writerow([])

                # Next Tier Info
                writer.writerow(["--- Next Tier Info ---"])
                writer.writerow(["Info Type", "Value"])
                for key, value in data["next_tier_info"].items():
                    writer.writerow([key.replace('_', ' ').title(), value])

            messagebox.showinfo("Export Success", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to CSV: {e}")

    def _export_to_pdf(self):
        if not FPDF:
            messagebox.showerror("Error", "fpdf library not found. Please install it using 'pip install fpdf'")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF files", "*.pdf")],
                                                title="Save Calculations to PDF")
        if not file_path:
            return

        try:
            data = self._get_all_calculated_data()
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Atlas Earth Calculator Report", 0, 1, "C")
            pdf.ln(10)

            def add_section(title, content_list):
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 10, title, 0, 1, "L")
                pdf.set_font("Arial", "", 10)
                for item in content_list:
                    pdf.cell(0, 7, item, 0, 1, "L")
                pdf.ln(5)

            user_inputs_content = []
            for key, value in data["user_inputs"].items():
                user_inputs_content.append(f"{key.replace('_', ' ').title()}: {value}")
            add_section("User Inputs", user_inputs_content)

            current_earnings_content = []
            current_earnings_content.append(f"{'Timeframe':<15} {'Base Earnings':<20} {'With Ad Boost':<20}")
            for tf, values in data["current_earnings"].items():
                base_val = values['base'] if isinstance(values['base'], str) else f"{values['base']:.10f}"
                boosted_val = values['boosted'] if isinstance(values['boosted'], str) else f"{values['boosted']:.10f}"
                current_earnings_content.append(f"{f'Per {tf.capitalize()}:':<15} {base_val:<20} {boosted_val:<20}")
            add_section("Current Earnings", current_earnings_content)

            goal_data_content = []
            for key, value in data["goal_data"].items():
                if isinstance(value, dict):
                    goal_data_content.append(f"{key.replace('_', ' ').title()}:")
                    for sub_key, sub_value in value.items():
                        goal_data_content.append(f"  {sub_key.replace('_', ' ').title()}: {sub_value}")
                else:
                    goal_data_content.append(f"{key.replace('_', ' ').title()}: {value}")
            add_section("Parcels for Goal", goal_data_content)

            next_tier_content = []
            for key, value in data["next_tier_info"].items():
                next_tier_content.append(f"{key.replace('_', ' ').title()}: {value}")
            add_section("Next Tier Info", next_tier_content)

            pdf.output(file_path)
            messagebox.showinfo("Export Success", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to PDF: {e}")


    def clear_all(self):
        """Resets all input fields to their default values."""
        if messagebox.askyesno("Clear All", "Are you sure you want to clear all values?"):
            for p_type in self.parcel_vars.keys():
                self.parcel_vars[p_type].set("0")
            self.badge_count_var.set("0")
            self.boost_hours_var.set("0") # Corrected default to "0"
            self.srb_boost_enabled.set(False)
            self.fictive_badge_boost_enabled.set(False)
            self.fictive_badge_boost_percent_var.set("0.0")
            self.selected_region_var.set("United States")
            self._toggle_fictive_badge_boost()
            self.update_all_calculations()


# --- Main execution block ---
if __name__ == "__main__":
    root = tk.Tk()
    app = AtlasEarthApp(root)
    root.mainloop()